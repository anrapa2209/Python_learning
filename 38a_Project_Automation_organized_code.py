import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)            #reusing the previous code to process 1000s of spreadsheets with different file names
    sheet = wb['Sheet1']
    #if we use this to automate the process of updating 1000's of spreadsheets, this prog wouldn't work
    #since this program only relies on the file transactions.xlsx
    #so, we need to reorganize this code & move it inside a function

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
             min_row=2,
             max_row=sheet.max_row,
             min_col=4,
             max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    wb.save(filename)           #override the above file with the updated price