#Excel spreadsheets: Program to process 1000s of spreadsheets in under a second
import openpyxl as xl
from openpyxl.chart import BarChart, Reference          #first letter of every word capitalized

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']                #make sure to use cap S, as it is case sensitive
cell = sheet['a1']                  #give the coordinate of the cell, combination of the column & row
cell = sheet.cell(1, 1)             #alternative approach to get a cell instead of using []
#print(cell.value)                  #printing value of the cell
#now we need to iterate over all the rows & we need to fill the values in the third column (price)
#print(sheet.max_row)               #to know how many rows we have in this spreadsheet
#So, we need to have a for loop that will generate the numbers 1 to 4

#for row in range(1, sheet.max_row + 1):     #adding 1, if sheet.max_row is alone given, range will go from 1 until sheet.max_row (excluding this value)
    #print(row)
#since the first row is header, we can ignore its values, so start from 2
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    #print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)        #limiting the range of cells we select, to the 4th column

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')             #specify the coordinate, where you want to add this chart

wb.save('transactions2.xlsx')           #new file 'transactions2' is created
#to add a chart, we need to import a couple of classes on the top
#then, select a range of values (for this eg., 4th column, corrected price is chosen)
#after for loop, we use the Reference class to select a range of values
#first argument in the constructor is 'Sheet', then we add 4 keyword arguments: min_row, max_row, min_col, max_col

#we organize our code by adding functions: see 38a